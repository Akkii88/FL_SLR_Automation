"""
Excel Export Service
=====================
Generates professional multi-sheet Excel exports.

Sheets:
- README
- Papers
- Search Log
- Provenance
- Screening
- Full Text Screening
- Claims
- Experiments
- Evidence Quality
- Ranking
- PRISMA
- Audit Log

Uses openpyxl for professional formatting:
- Frozen headers
- Autofilter
- Readable column widths
- Wrapped text
- Consistent date formats
"""

import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.core.config import settings
from app.models.paper import Paper
from app.models.search_run import SearchRun, SourceProvenance
from app.models.screening import ScreeningDecision, AuditLog
from app.models.deduplication import DeduplicationLog
from app.models.extraction import Claim, Experiment, Condition, EvidenceQuality
from app.models.pdf_file import PdfFile, PaperNote

logger = logging.getLogger(__name__)

# Styles
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_COLOR = Font(bold=True, size=11, color="FFFFFF")
WRAPPED_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
DATE_FORMAT = "YYYY-MM-DD HH:MM:SS"
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


class ExcelExportService:
    """
    Generates professional multi-sheet Excel exports.
    """

    def __init__(self, db: Session):
        self.db = db
        self.wb = Workbook()

    def generate_full_export(self, output_path: Optional[Path] = None) -> Path:
        """
        Generate the complete Excel export with all sheets.
        
        Args:
            output_path: Optional custom output path. Defaults to data/exports/
            
        Returns:
            Path to the generated Excel file
        """
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = settings.project_root / "data" / "exports" / f"fl_slr_export_{timestamp}.xlsx"

        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Remove default sheet
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])

        # Generate all sheets
        self._create_readme_sheet()
        self._create_papers_sheet()
        self._create_search_log_sheet()
        self._create_provenance_sheet()
        self._create_screening_sheet()
        self._create_claims_sheet()
        self._create_experiments_sheet()
        self._create_evidence_quality_sheet()
        self._create_prisma_sheet()
        self._create_audit_log_sheet()

        self.wb.save(output_path)
        logger.info(f"Excel export saved to: {output_path}")
        return output_path

    def _setup_sheet(self, title: str, headers: list[str]):
        """Create a sheet with formatted headers."""
        ws = self.wb.create_sheet(title=title)

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT_COLOR
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

        # Freeze top row
        ws.freeze_panes = "A2"

        # Auto-filter
        if len(headers) > 0:
            last_col = get_column_letter(len(headers))
            ws.auto_filter.ref = f"A1:{last_col}1"

        return ws

    def _auto_width(self, ws):
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = max(adjusted_width, 12)

    def _create_readme_sheet(self):
        """Create the README sheet with project information."""
        ws = self.wb.create_sheet("README", 0)

        info = [
            ["FL-SLR Automation", ""],
            ["Project", 'Is "Best" Really Best? A Systematic Review of Evidence Quality Behind Federated Learning Algorithm Superiority Claims under Non-IID Data'],
            ["Export Date", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")],
            ["Software Version", "1.0.0"],
            ["", ""],
            ["This workbook contains the following sheets:", ""],
            ["Papers", "All bibliographic records with metadata and screening status"],
            ["Search Log", "Complete search history with queries and results"],
            ["Provenance", "Source provenance for each paper"],
            ["Screening", "Title/abstract screening decisions"],
            ["Claims", "Extracted comparative claims"],
            ["Experiments", "Experimental setups for each claim"],
            ["Evidence Quality", "5-dimension evidence quality assessments"],
            ["PRISMA", "PRISMA flow counts"],
            ["Audit Log", "Complete audit trail"],
        ]

        for row_idx, (key, value) in enumerate(info, 1):
            ws.cell(row=row_idx, column=1, value=key).font = Font(bold=True, size=12 if row_idx <= 4 else 10)
            ws.cell(row=row_idx, column=2, value=value)

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 80

    def _create_papers_sheet(self):
        """Create the Papers sheet."""
        headers = [
            "ID", "Title", "Authors", "Year", "Source", "DOI", "OpenAlex ID",
            "Citations", "OA Status", "PDF URL", "Screening Status",
            "Screening Decision", "Exclusion Reason", "Duplicate Status",
            "Duplicate Of", "Search Families", "Created At"
        ]
        ws = self._setup_sheet("Papers", headers)

        papers = self.db.query(Paper).order_by(Paper.id).all()

        for row_idx, paper in enumerate(papers, 2):
            families = ", ".join(set(
                p.search_family for p in paper.provenance if p.search_family
            ))

            values = [
                paper.id,
                paper.title,
                paper.authors,
                paper.publication_year,
                paper.source,
                paper.doi,
                paper.openalex_id,
                paper.citation_count,
                paper.oa_status,
                paper.pdf_url,
                paper.screening_status,
                paper.screening_decision,
                paper.exclusion_reason,
                paper.duplicate_status,
                paper.duplicate_of,
                families,
                paper.created_at.strftime(DATE_FORMAT) if paper.created_at else "",
            ]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = WRAPPED_ALIGNMENT
                cell.border = THIN_BORDER

        self._auto_width(ws)

    def _create_search_log_sheet(self):
        """Create the Search Log sheet."""
        headers = [
            "Search ID", "Source", "Family", "Query", "Date", "Start Time",
            "End Time", "Year Filter", "Results Retrieved", "Records Saved",
            "Pages", "Duration (s)", "Retries", "Errors", "Notes"
        ]
        ws = self._setup_sheet("Search Log", headers)

        runs = self.db.query(SearchRun).order_by(SearchRun.id).all()

        for row_idx, run in enumerate(runs, 2):
            values = [
                run.id,
                run.source,
                run.search_family,
                run.exact_query,
                run.search_date.strftime(DATE_FORMAT) if run.search_date else "",
                run.start_time.strftime(DATE_FORMAT) if run.start_time else "",
                run.end_time.strftime(DATE_FORMAT) if run.end_time else "",
                run.year_filter,
                run.results_retrieved,
                run.records_saved,
                run.pages_retrieved,
                run.duration_seconds,
                run.retries,
                run.errors,
                run.notes,
            ]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = WRAPPED_ALIGNMENT
                cell.border = THIN_BORDER

        self._auto_width(ws)

    def _create_provenance_sheet(self):
        """Create the Provenance sheet."""
        headers = ["ID", "Paper ID", "Source", "Search Family", "Retrieval Timestamp"]
        ws = self._setup_sheet("Provenance", headers)

        provenance = self.db.query(SourceProvenance).order_by(SourceProvenance.id).all()

        for row_idx, prov in enumerate(provenance, 2):
            values = [
                prov.id,
                prov.paper_id,
                prov.source,
                prov.search_family,
                prov.retrieval_timestamp.strftime(DATE_FORMAT) if prov.retrieval_timestamp else "",
            ]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = WRAPPED_ALIGNMENT
                cell.border = THIN_BORDER

        self._auto_width(ws)

    def _create_screening_sheet(self):
        """Create the Screening sheet."""
        headers = [
            "Decision ID", "Paper ID", "Stage", "Q1 FL Comparison",
            "Q2 Non-IID", "Q3 Superiority Claim", "Q4 Full Text Available",
            "Decision", "Exclusion Reason", "Notes", "Decided By", "Created At"
        ]
        ws = self._setup_sheet("Screening", headers)

        decisions = self.db.query(ScreeningDecision).order_by(ScreeningDecision.id).all()

        for row_idx, dec in enumerate(decisions, 2):
            values = [
                dec.id,
                dec.paper_id,
                dec.stage,
                dec.q1_fl_comparison,
                dec.q2_non_iid,
                dec.q3_superiority_claim,
                dec.q4_full_text_available,
                dec.decision,
                dec.exclusion_reason,
                dec.notes,
                dec.decided_by,
                dec.created_at.strftime(DATE_FORMAT) if dec.created_at else "",
            ]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = WRAPPED_ALIGNMENT
                cell.border = THIN_BORDER

        self._auto_width(ws)

    def _create_claims_sheet(self):
        """Create the Claims sheet."""
        headers = [
            "Claim ID", "Paper ID", "Claim Text", "Claim Scope",
            "Algorithms Compared", "Winner", "Datasets", "Non-IID Type",
            "Partition Method", "Heterogeneity Param", "Evidence Page",
            "Evidence Section", "Evidence Table", "Evidence Figure",
            "Evidence Snippet", "Notes"
        ]
        ws = self._setup_sheet("Claims", headers)

        claims = self.db.query(Claim).order_by(Claim.id).all()

        for row_idx, claim in enumerate(claims, 2):
            values = [
                claim.id,
                claim.paper_id,
                claim.claim_text,
                claim.claim_scope,
                claim.algorithms_compared,
                claim.winner_algorithm,
                claim.datasets,
                claim.non_iid_type,
                claim.partition_method,
                claim.heterogeneity_param,
                claim.evidence_page,
                claim.evidence_section,
                claim.evidence_table,
                claim.evidence_figure,
                claim.evidence_snippet,
                claim.notes,
            ]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = WRAPPED_ALIGNMENT
                cell.border = THIN_BORDER

        self._auto_width(ws)

    def _create_experiments_sheet(self):
        """Create the Experiments sheet."""
        headers = [
            "Experiment ID", "Claim ID", "Name", "Dataset",
            "Non-IID Type", "Partition Method", "Heterogeneity Param",
            "Independent Runs", "Seed Reported",
            "Condition ID", "Algorithm", "Metric", "Metric Value",
            "Ranking", "Is Winner", "SD", "CI"
        ]
        ws = self._setup_sheet("Experiments", headers)

        experiments = self.db.query(Experiment).order_by(Experiment.id).all()

        row_idx = 2
        for exp in experiments:
            for cond in exp.conditions:
                values = [
                    exp.id,
                    exp.claim_id,
                    exp.experiment_name,
                    exp.dataset,
                    exp.non_iid_type,
                    exp.partition_method,
                    exp.heterogeneity_param,
                    exp.independent_runs,
                    exp.random_seed_reported,
                    cond.id,
                    cond.algorithm,
                    cond.metric_name,
                    cond.metric_value,
                    cond.ranking_position,
                    cond.is_winner,
                    cond.standard_deviation,
                    cond.confidence_interval,
                ]

                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = WRAPPED_ALIGNMENT
                    cell.border = THIN_BORDER

                row_idx += 1

        self._auto_width(ws)

    def _create_evidence_quality_sheet(self):
        """Create the Evidence Quality sheet."""
        headers = [
            "EQ ID", "Claim ID", "Independent Runs", "Seed Reported",
            "Uncertainty", "SD Type", "CI Level",
            "Direct Stats Test", "Mechanism Test", "Statistical Unit",
            "Effect Size", "Effect Value",
            "Matched Partition", "HP Tuning Fairness",
            "Ranking Robustness", "Evidence Basis",
            "Author Claim vs Evidence", "Limitation", "Code Repo",
            "Notes"
        ]
        ws = self._setup_sheet("Evidence Quality", headers)

        eq_records = self.db.query(EvidenceQuality).order_by(EvidenceQuality.id).all()

        for row_idx, eq in enumerate(eq_records, 2):
            values = [
                eq.id,
                eq.claim_id,
                eq.independent_runs,
                eq.random_seed_reported,
                eq.uncertainty_reporting,
                eq.sd_type,
                eq.ci_level,
                eq.direct_statistical_test,
                eq.mechanism_level_statistical_test,
                eq.statistical_unit,
                eq.effect_size_reported,
                eq.effect_size_value,
                eq.matched_client_partition,
                eq.hyperparameter_tuning_fairness,
                eq.ranking_robustness,
                eq.evidence_basis,
                eq.author_claim_vs_evidence,
                eq.important_limitation,
                eq.code_repository_url,
                eq.notes,
            ]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = WRAPPED_ALIGNMENT
                cell.border = THIN_BORDER

        self._auto_width(ws)

    def _create_prisma_sheet(self):
        """Create the PRISMA sheet."""
        ws = self.wb.create_sheet("PRISMA")

        from app.services.prisma import PrismaService
        prisma = PrismaService(self.db)
        flow = prisma.get_prisma_flow()

        headers = ["Stage", "Metric", "Value"]
        ws.cell(row=1, column=1, value="Stage").font = HEADER_FONT_COLOR
        ws.cell(row=1, column=1).fill = HEADER_FILL
        ws.cell(row=1, column=2, value="Metric").font = HEADER_FONT_COLOR
        ws.cell(row=1, column=2).fill = HEADER_FILL
        ws.cell(row=1, column=3, value="Value").font = HEADER_FONT_COLOR
        ws.cell(row=1, column=3).fill = HEADER_FILL

        row_idx = 2

        # Identification
        ws.cell(row=row_idx, column=1, value="Identification")
        ws.cell(row=row_idx, column=2, value="Total Records Retrieved")
        ws.cell(row=row_idx, column=3, value=flow["identification"]["total_records_retrieved"])
        row_idx += 1

        # Screening
        ws.cell(row=row_idx, column=1, value="Screening")
        ws.cell(row=row_idx, column=2, value="Duplicates Removed")
        ws.cell(row=row_idx, column=3, value=flow["screening"]["duplicates_removed"])
        row_idx += 1

        ws.cell(row=row_idx, column=1, value="Screening")
        ws.cell(row=row_idx, column=2, value="Unique Records")
        ws.cell(row=row_idx, column=3, value=flow["screening"]["unique_records"])
        row_idx += 1

        ws.cell(row=row_idx, column=1, value="Screening")
        ws.cell(row=row_idx, column=2, value="Records Screened")
        ws.cell(row=row_idx, column=3, value=flow["screening"]["records_screened"])
        row_idx += 1

        ws.cell(row=row_idx, column=1, value="Screening")
        ws.cell(row=row_idx, column=2, value="Title/Abstract Excluded")
        ws.cell(row=row_idx, column=3, value=flow["screening"]["title_abstract_excluded"])
        row_idx += 1

        # Included
        ws.cell(row=row_idx, column=1, value="Included")
        ws.cell(row=row_idx, column=2, value="Final Included")
        ws.cell(row=row_idx, column=3, value=flow["included"]["final_included"])
        row_idx += 1

        ws.cell(row=row_idx, column=1, value="Included")
        ws.cell(row=row_idx, column=2, value="Papers with Claims")
        ws.cell(row=row_idx, column=3, value=flow["included"]["papers_with_claims"])

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 15

    def _create_audit_log_sheet(self):
        """Create the Audit Log sheet."""
        headers = [
            "ID", "Timestamp", "Action", "Entity Type", "Entity ID",
            "Description", "Old Value", "New Value", "Actor", "Paper ID"
        ]
        ws = self._setup_sheet("Audit Log", headers)

        logs = self.db.query(AuditLog).order_by(AuditLog.timestamp).all()

        for row_idx, log in enumerate(logs, 2):
            values = [
                log.id,
                log.timestamp.strftime(DATE_FORMAT) if log.timestamp else "",
                log.action,
                log.entity_type,
                log.entity_id,
                log.description,
                log.old_value,
                log.new_value,
                log.actor,
                log.paper_id,
            ]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = WRAPPED_ALIGNMENT
                cell.border = THIN_BORDER

        self._auto_width(ws)
